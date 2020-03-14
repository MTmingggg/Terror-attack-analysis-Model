import xlrd
import xlwt
import csv
import pandas as pd
import numpy as np
import scipy as sp
from sklearn.cluster import KMeans
import matplotlib.pyplot as plt
import re


'''
# 数据格式转换
f = open('E:/data.csv', 'w', newline='', encoding='utf8')
writer = csv.writer(f)
writer.writerow(['eventid', 'iyear', 'imonth', 'iday', 'approxdate', 'extended', 'resolution', 'country',
                 'country_txt', 'region', 'region_txt', 'provstate', 'city', 'latitude', 'longitude',
                 'specificity', 'vicinity', 'location', 'summary', 'crit1', 'crit2', 'crit3', 'doubtterr',
                 'alternative', 'alternative_txt', 'multiple', 'success', 'suicide', 'attacktype1',
                 'attacktype1_txt', 'attacktype2', 'attacktype2_txt', 'attacktype3', 'attacktype3_txt',
                 'targtype1', 'targtype1_txt', 'targsubtype1', 'targsubtype1_txt', 'corp1', 'target1',
                 'natlty1', 'natlty1_txt', 'targtype2', 'targtype2_txt', 'targsubtype2', 'targsubtype2_txt',
                 'corp2', 'target2', 'natlty2', 'natlty2_txt', 'targtype3', 'targtype3_txt', 'targsubtype3',
                 'targsubtype3_txt', 'corp3', 'target3', 'natlty3', 'natlty3_txt', 'gname', 'gsubname', 'gname2',
                 'gsubname2', 'gname3', 'gsubname3', 'motive', 'guncertain1', 'guncertain2', 'guncertain3',
                 'individual', 'nperps', 'nperpcap', 'claimed', 'claimmode', 'claimmode_txt', 'claim2',
                 'claimmode2', 'claimmode2_txt', 'claim3', 'claimmode3', 'claimmode3_txt', 'compclaim',
                 'weaptype1', 'weaptype1_txt', 'weapsubtype1', 'weapsubtype1_txt', 'weaptype2', 'weaptype2_txt',
                 'weapsubtype2', 'weapsubtype2_txt', 'weaptype3', 'weaptype3_txt', 'weapsubtype3',
                 'weapsubtype3_txt', 'weaptype4', 'weaptype4_txt', 'weapsubtype4', 'weapsubtype4_txt',
                 'weapdetail', 'nkill', 'nkillus', 'nkillter', 'nwound', 'nwoundus', 'nwoundte', 'property',
                 'propextent', 'propextent_txt', 'propvalue', 'propcomment', 'ishostkid', 'nhostkid',
                 'nhostkidus', 'nhours', 'ndays', 'divert', 'kidhijcountry', 'ransom', 'ransomamt',
                 'ransomamtus', 'ransompaid', 'ransompaidus', 'ransomnote', 'hostkidoutcome', 'hostkidoutcome_txt',
                 'nreleased', 'addnotes', 'scite1', 'scite2', 'scite3', 'dbsource', 'INT_LOG', 'INT_IDEO',
                 'INT_MISC', 'INT_ANY', 'related'])
data = xlrd.open_workbook('E:/附件1.xlsx')
table = data.sheets()[0]
nrows = table.nrows
print(nrows)
for i in range(nrows):
    if i == 0:
        continue
    writer.writerow(table.row_values(i)[:])
'''
'''
# 剔除缺失值较多的字段
df = pd.read_csv(open('E:/data.csv', encoding='utf8'))
df2 = df[['eventid', 'iyear', 'imonth', 'iday','extended','country','country_txt', 'region', 'region_txt',
          'provstate', 'city', 'latitude', 'longitude','summary', 'crit1', 'crit2', 'crit3', 'doubtterr',
          'multiple', 'success', 'suicide', 'attacktype1','attacktype1_txt', 'attacktype2', 'attacktype2_txt',
          'targtype1', 'targtype1_txt', 'targsubtype1', 'targsubtype1_txt','corp1', 'target1','natlty1',
          'natlty1_txt','targtype2', 'targtype2_txt', 'targsubtype2', 'targsubtype2_txt', 'corp2', 'target2',
          'natlty2', 'natlty2_txt','gname', 'gsubname', 'gname2','gsubname2','motive', 'guncertain1','claimed',
          'claimmode', 'claimmode_txt', 'claim2', 'claimmode2', 'claimmode2_txt','weaptype1', 'weaptype1_txt',
          'weapsubtype1', 'weapsubtype1_txt', 'weaptype2', 'weaptype2_txt','weapsubtype2', 'weapsubtype2_txt',
          'weapdetail', 'nkill','nkillter','nwound','nwoundte','property','propextent', 'propextent_txt', 'ishostkid', 'nhostkid',
          'nhours', 'ndays', 'INT_LOG', 'INT_IDEO', 'INT_MISC', 'INT_ANY', 'related']]
f = open('E:/data2.csv', 'w', encoding='utf8', newline='')
writer = csv.writer(f)
writer.writerow(['eventid', 'iyear', 'imonth', 'iday','extended','country','country_txt', 'region', 'region_txt',
          'provstate', 'city', 'latitude', 'longitude','summary', 'crit1', 'crit2', 'crit3', 'doubtterr',
          'multiple', 'success', 'suicide', 'attacktype1','attacktype1_txt', 'attacktype2', 'attacktype2_txt',
          'targtype1', 'targtype1_txt', 'targsubtype1', 'targsubtype1_txt','corp1', 'target1','natlty1',
          'natlty1_txt','targtype2', 'targtype2_txt', 'targsubtype2', 'targsubtype2_txt', 'corp2', 'target2',
          'natlty2', 'natlty2_txt','gname', 'gsubname', 'gname2','gsubname2','motive', 'guncertain1','claimed',
          'claimmode', 'claimmode_txt', 'claim2', 'claimmode2', 'claimmode2_txt','weaptype1', 'weaptype1_txt',
          'weapsubtype1', 'weapsubtype1_txt', 'weaptype2', 'weaptype2_txt','weapsubtype2', 'weapsubtype2_txt',
          'weapdetail', 'nkill','nkillter','nwound','nwoundte','property','propextent', 'propextent_txt', 'ishostkid', 'nhostkid',
          'nhours', 'ndays', 'INT_LOG', 'INT_IDEO', 'INT_MISC', 'INT_ANY', 'related'])
for i in range(len(df2)):
    temp = df2.loc[i][:]
    temp = temp.tolist()
    writer.writerow(temp)
'''
'''
# 筛选部分字段，进行数据预处理
# 数值化
df = pd.read_csv(open('E:/data2.csv', encoding='utf8')) # 读取数据
df.replace(-9, 0.5, inplace=True) # 用0.5替换数据中的-9
df['city'].fillna('n',inplace=True)  # 用字符n填充city字段的缺失值，以便于后面处理
df.fillna(0, inplace=True)  # 用0填充数据缺失值
region = df['region'].tolist()  # 提取region字段，并生成列表
attacktype = df['attacktype1'].tolist()  # 提取attacktype1字段，并生成列表
targtype = df['targtype1'].tolist()  # 提取targtype1字段，并生成列表
weapontype = df['weaptype1'].tolist()  # 提取weaptype1字段，并生成列表
nkill = df['nkill'].tolist()  # 提取nkill字段，并生成列表
nkillter = df['nkillter'].tolist()  # 提取nkillter字段，并生成列表
nwound = df['nwound'].tolist()  # 提取nwound字段，并生成列表
nwoundte = df['nwoundte'].tolist()  # 提取nwoundte字段，并生成列表

# 地区数值化
region_new = []
for i in range(len(region)):
    if int(region[i]) in [1, 4, 8, 9, 12]:  # 将北美、东亚、西欧、东欧、澳大利亚和大洋洲替换为数值3
        region_new.append(3)
    elif int(region[i]) in [2, 3, 5, 7]:  # 将中美洲和加勒比海地区、南美、东南亚、中亚设为2
        region_new.append(2)
    else:
        region_new.append(1)  # 与地区设置为1

# 国家数值化
# 国家划分为发达国家，发展中国家以及其他国家
fa = ['卢森堡','挪威','瑞士','爱尔兰','丹麦','冰岛','瑞典','大不列颠联合王国','英国','奥地利','荷兰','芬兰',
      '比利时','法国','德国','意大利','西班牙','希腊','葡萄牙','美国','加拿大','日本','新加坡','澳大利亚',
      '新西兰','塞浦路斯','巴哈马','斯洛文尼亚','以色列','韩国','马耳他','匈牙利','捷克','波兰','斯洛伐克',
      '安道尔','巴林','巴巴多斯','文莱','爱沙尼亚','中国香港','列支敦士登','摩纳哥','卡塔尔','圣马力诺','阿联酋']
df_con = pd.read_csv(open('E:/建模数据/country.csv', encoding='utf8'))   # 读取国家数据，通过辅助文件调用百度翻译api将
country = df_con['country'].tolist()  # 国家名翻译为中文
f_sec_country = open('E:/建模数据/发展中国家.txt', encoding='utf8')
country_sec = []  # 保存发展中国家数据
for line in f_sec_country.readlines():
    country_sec.append(line.strip())
country_new = []
for i in range(len(country)):
    if country[i] in fa:
        country_new.append(3)  # 如果是发达国家，设置为3
    elif country[i] in country_sec:
        country_new.append(2)  # 如果是发展中国家，设置为2
    else:
        country_new.append(1)  # 其余设置为1

# 城市数值化
# 将国家首都设置为3，除首都外的国际大城市设置为2，其余设置为1
cc = xlrd.open_workbook('E:/建模数据/国家名称包含首都名称.xls')  # 读取首都数据
table = cc.sheets()[0] # 读取第一页
nrows = table.nrows
data = []  # 保存Excel数据
for i in range(nrows):
    if i == 0:
        continue
    temp = table.row_values(i)[:]
    data.append(temp)
capital = []  # 获取首都数据
for i in range(len(data)):
    capital.append(data[i][-1])

df_city = pd.read_csv(open('E:/city.csv', encoding='utf8'))  # 读取城市数据，利用辅助文件调用百度翻译api将
city = df_city['city'].tolist()  # 城市名翻译为中文
f_sec = open('E:/著名城市.txt', encoding='utf8')  # 读取除首都外的大城市名
city_sec = []  # 保存大城市名
for line in f_sec.readlines():
    city_sec.append(line.strip())
city_new = []
for i in range(len(city)):
    if city[i] in capital:  # 如果是首都，设置为3
        city_new.append(3)
    elif city[i] in city_sec:
        city_new.append(2)  # 如果是大城市，设置为2
    else:
        city_new.append(1) # 其余的设置为1

# 攻击类型数值化
# 根据攻击类型出现频率设置
attacktype_new = []
for i in range(len(attacktype)):
    attacktype_new.append(attacktype.count(attacktype[i])/len(attacktype))

# 目标数值化
# 根据目标类型出现频率设置
targtype_new = []
for i in range(len(targtype)):
    targtype_new.append(targtype.count(targtype[i])/len(targtype))

# 武器数值化
# 根据武器类型出现频率设置
weapontype_new = []
for i in range(len(weapontype)):
    weapontype_new.append(weapontype.count(weapontype[i])/len(weapontype))

# 伤亡人数数值化
# 剔除恐怖分子伤亡人数，仅保留平民伤亡人数
kill = []  # 保存平民死亡人数
wound = []  # 保存平民受伤人数
for i in range(len(nkill)):  
    kill.append(nkill[i]-nkillter[i])
    wound.append(nwound[i]-nwoundte[i])
f_train = open('E:/train2.csv', 'w', encoding='utf8', newline='')  # 保存设置字段数据
writer = csv.writer(f_train)
writer.writerow(['extended','success', 'suicide','propextent','country','region','city','attacktype','targtype',
                 'weapontype','kill','wound'])
df2 = df[['extended','success', 'suicide','propextent']]
for i in range(len(df2)):
    temp = df2.loc[i][:]
    temp = temp.tolist()
    temp += [country_new[i], region_new[i], city_new[i], attacktype_new[i], targtype_new[i], weapontype_new[i],kill[i],wound[i]]
    writer.writerow(temp)
'''
'''
# 用于将设置字段存入excel文件
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

myexcel = Workbook()
mysheet = myexcel.create_sheet('test')
csvfile = open('E:/train.csv', encoding='utf8')
reader = csv.reader(csvfile)
row = 1
title_list = ['extended','latitude', 'longitude','success', 'suicide','nkill',
          'nwound','propextent','country','region','city','attacktype','targtype','weapontype']
for i in range(1,len(title_list)+1):
    mysheet.cell(row = row, column = i).value = title_list[i-1]

for line in reader:
    if line == ['extended','latitude', 'longitude','success', 'suicide','nkill',
          'nwound','propextent','country','region','city','attacktype','targtype','weapontype']:
        continue
    r = 1
    for i in line:
        print(row,r)
        mysheet.cell(row=row+1, column=r).value = float(i)
        r += 1
    row += 1

myexcel.save("E:/myexcel.xlsx")
'''
'''
# 聚类
df = pd.read_csv(open('E:/train2.csv', encoding='utf8'))
df_norm = (df - df.min()) / (df.max() - df.min())  # 数据归一化
data = []  # 保存数据
for i in range(len(df_norm)):
    temp = df.loc[i][:]
    temp = temp.tolist()
    data.append(temp)

dataset = np.array(data)  # 转为矩阵
kmeans = KMeans(n_clusters=5, random_state=0, max_iter=300).fit(dataset)  # 调用kmeans聚类
labels = kmeans.labels_  # 聚类标签
centroids = kmeans.cluster_centers_  # 聚类质心
print("kmeans down")
print('0类质心：', centroids[0])
print('1类质心：', centroids[1])
print('2类质心：', centroids[2])
print('3类质心：', centroids[3])
print('4类质心：', centroids[4])
print(centroids)
r1 = pd.Series(kmeans.labels_).value_counts()  # 统计各个类别的数目
r2 = pd.DataFrame(kmeans.cluster_centers_)  # 找出聚类中心
r = pd.concat([r2, r1], axis = 1)  # 横向连接（0是纵向），得到聚类中心对应的类别下的数目
r.columns = ['extended', 'success', 'suicide', 'propextent', 'country', 'region', 'city', 'attacktype',
             'targtype', 'weapontype', 'kill', 'wound']\
            + [u'类别数目']  # 重命名表头
# 下面几行代码是为了在聚类结果中加入id，运行以下被注释代码时，修改如下：
# dfs = pd.DataFrame(data2, index=labels,columns=['extended', 'success', 'suicide', 'propextent', 'country',
#                                  'region', 'city', 'attacktype', 'targtype', 'weapontype', 'kill', 'wound'])
# dfs.to_excel('E:/cluster3.xlsx')
# df_data2 = pd.read_csv(open('E:/data2.csv', encoding='utf8'))
# eventid = df_data2['eventid'].tolist()
# data2 = []
# for i in range(len(eventid)):
#     data2.append([eventid[i]]+data[i])
dfs = pd.DataFrame(dataset, index=labels, columns=['extended', 'success', 'suicide', 'propextent', 'country',
                            'region', 'city', 'attacktype', 'targtype', 'weapontype', 'kill', 'wound'])
dfs.to_excel('E:/cluster2.xlsx')  # 将聚类后结果保存入excel文件
print(r)
'''
'''
# 评价
model = re.compile(r'[.\d]+')  # 正则表达式model
f_quan = open('E:/权重.txt', encoding='utf8')  # 打开保存权重文件
factor = []  # 保存权重数据
for line in f_quan.readlines():
    factor.append(float(''.join(model.findall(line.strip()))))
factor = np.array(factor)
data = xlrd.open_workbook('E:/cluster2.xlsx')  # 打开聚类结果
table = data.sheets()[0]
nrows = table.nrows
score = []
label = []
for i in range(nrows):
    if i == 0:
        continue
    temp = table.row_values(i)[:]
    label.append(int(temp[0]))
    tt = np.array(temp[1:])
    ss = tt*factor
    score.append([int(temp[0]), sum(ss)])
# 以下是危害性分级
len_0 = label.count(0)
len_1 = label.count(1)
len_2 = label.count(2)
len_3 = label.count(3)
len_4 = label.count(4)
sum_0 = 0
sum_1 = 0
sum_2 = 0
sum_3 = 0
sum_4 = 0
for i in range(len(score)):
    if score[i][0] == 0:
        sum_0 += score[i][1]
    elif score[i][0] == 1:
        sum_1 += score[i][1]
    elif score[i][0] == 2:
        sum_2 += score[i][1]
    elif score[i][0] == 3:
        sum_3 += score[i][1]
    elif score[i][0] == 4:
        sum_4 += score[i][1]
print("0类：", sum_0/len_0)
print("1类：", sum_1/len_1)
print("2类：", sum_2/len_2)
print("3类：", sum_3/len_3)
print("4类：", sum_4/len_4)
'''
'''
# 选出十大恐袭
data = xlrd.open_workbook('E:/cluster3.xlsx')
table = data.sheets()[0]
nrows = table.nrows
id = []
tt = []
for i in range(nrows):
    if i == 0:
        continue
    temp = table.row_values(i)[:]
    if int(temp[0]) == 1:
        id.append(int(temp[1]))
    if int(temp[0]) == 2:
        id.append(int(temp[1]))
    if int(temp[0]) == 4:
        tt.append(temp)
model = re.compile(r'[.\d]+')
f_quan = open('E:/权重.txt', encoding='utf8')
factor = []
for line in f_quan.readlines():
    factor.append(float(''.join(model.findall(line.strip()))))
factor = np.array(factor)
score = {}
for i in range(len(tt)):
    ss = np.array(tt[i][2:])
    score[tt[i][1]] = sum(ss*factor)
pp = sorted(score.items(), key=lambda x: x[1], reverse=True)
num = 0
for i in range(len(pp)):
    if num < 7:
        id.append(pp[i][0])
        num += 1
print("近二十年来危害程度最高的十大恐袭事件id：")
for i in range(len(id)):
    print(int(id[i]))
'''
'''
# 将聚类号替换为危害性等级，并将结果存入excel文件
data = xlrd.open_workbook('E:/cluster3.xlsx')
table = data.sheets()[0]
nrows = table.nrows
tt = [] # 全部内容
for i in range(nrows):
    if i == 0:
        continue
    temp = table.row_values(i)[:]
    tt.append(temp)
ss = [] # 保存除聚类号外的内容
label = [] # 保存危害性等级数据
for i in range(len(tt)):
    if int(tt[i][0]) == 1:
        label.append(5)
        ss.append(tt[i][1:])
    elif int(tt[i][0]) == 2:
        label.append(4)
        ss.append(tt[i][1:])
    elif int(tt[i][0]) == 4:
        label.append(3)
        ss.append(tt[i][1:])
    elif int(tt[i][0]) == 3:
        label.append(2)
        ss.append(tt[i][1:])
    elif int(tt[i][0]) == 0:
        label.append(1)
        ss.append(tt[i][1:])
df = pd.DataFrame(ss, index=label, columns=['eventid','extended','success', 'suicide','propextent','country',
                                            'region','city','attacktype','targtype','weapontype','kill','wound'])
df.to_excel('E:/result_class.xlsx')
'''
'''
# 完成表1
data = xlrd.open_workbook('E:/result_class.xlsx')
table = data.sheets()[0]
nrows = table.nrows
tt = [] # 全部内容
for i in range(nrows):
    if i == 0:
        continue
    temp = table.row_values(i)[:]
    tt.append(temp)
l = [200108110012,200511180002,200901170021,201402110015,201405010071,201411070002,201412160041,201508010015,201705080012]
r = []
for i in range(len(tt)):
    if int(tt[i][1]) in l:
        r.append(int(tt[i][0]))
print(' '*2,"表1 典型事件危害级别")
print("事件编号",' '*7,"危害级别")
for i in range(len(r)):
    print(l[i],' '*6, r[i])
'''